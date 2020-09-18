from pathlib import Path
from openpyxl import Workbook, load_workbook
from rule_classes import ObjectGenerator, IntegerGenerator, StringGenerator, BooleanGenerator, DateGenerator, \
    Nesting, UUIDGenerator


class AnotherSubDetails(ObjectGenerator):
    polo = IntegerGenerator(start=1000, end=20000)


class SubDetails(ObjectGenerator):
    id = UUIDGenerator()
    yolo = IntegerGenerator(start=1000, end=20000)
    another_details = Nesting(relation_with=AnotherSubDetails)


class Details(ObjectGenerator):
    id = UUIDGenerator()
    money = IntegerGenerator(start=1000, end=20000)
    sub_details = Nesting(relation_with=SubDetails, many=True, many_count=2)


class Person(ObjectGenerator):
    id = UUIDGenerator()
    age = IntegerGenerator(start=1, end=10)
    phone_no = IntegerGenerator(start=9000000010, end=9999999999)
    eligible = BooleanGenerator()
    name = StringGenerator("country")
    date_time = DateGenerator()
    details = Nesting(relation_with=Details)


class SimpleObj(ObjectGenerator):
    id = UUIDGenerator()
    age = IntegerGenerator(start=1, end=10)
    phone_no = IntegerGenerator(start=9000000010, end=9999999999)
    eligible = BooleanGenerator()
    details = Nesting(relation_with=AnotherSubDetails)


class RandomDataGenerator:

    def __int__(self):
        pass

    def _generate_data(self, obj, many=False, many_count=0, relate_id=None, relate_name=None):
        if many:
            temp_list = list()
            for idx in range(many_count):
                temp_list.append(next(self._generate_data(obj, relate_id=relate_id, relate_name=relate_name)))
            yield temp_list
        else:
            temp_dict = dict()
            attr_to_use = {key: value for key, value in vars(obj).items() if not key.startswith('_')}
            for attr_name, attr_obj in attr_to_use.items():
                if isinstance(attr_obj, Nesting):
                    temp_dict[attr_name] = next(self._generate_data(attr_obj.relation_with, many=attr_obj.many,
                                                                    many_count=attr_obj.many_count,
                                                                    relate_name=obj.__name__,
                                                                    relate_id=temp_dict.get(attr_obj.relate_by)))
                else:
                    temp_dict[attr_name] = attr_obj.get_value()

            if relate_id and relate_name:
                temp_dict[relate_name.lower() + '_id'] = relate_id

            yield temp_dict

    def write_csv(self, obj_to_create, obj_name, folder_path):
        excel_name = folder_path.joinpath(obj_name + '.xlsx')
        if excel_name.exists():
            wb = load_workbook(filename=excel_name.__str__())
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([key for key, value in obj_to_create.items() if not isinstance(value, (dict, list))])

        values_to_write = list()
        for attr_name, attr_value in obj_to_create.items():
            if isinstance(attr_value, dict):
                self.write_csv(attr_value, attr_name, folder_path)
            elif isinstance(attr_value, list):
                for obj in attr_value:
                    self.write_csv(obj, attr_name, folder_path)
            else:
                values_to_write.append(attr_value)
        ws.append(values_to_write)
        wb.save(filename=excel_name.__str__())

    def generate_csv(self, obj_to_create, folder_path=Path(r"F:\python_projects\random_data_genrator\src")):
        obj_to_create._validate()
        for i in range(5):
            # for data in self._generate_data(obj_to_create):
            print(next(self._generate_data(obj_to_create)))
            # self.write_csv(data, obj_to_create.__name__, folder_path)

import pprint

data_generator_ins = RandomDataGenerator()
data_generator_ins.generate_csv(Person)
#pp = pprint.PrettyPrinter(indent=4)

#pp.pprint(next(data_generator_ins._generate_data(Person)))
